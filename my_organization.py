import openpyxl
import calendar

print('Hello! Welcome in my program!')

while True:
    print('Choose what do you want to make:')
    print('1. Calendar and adding tasks to different days.')
    print('2. Incomes and expenses.')
    print('3. Exit.')
    start_choice = int(input(('Hello! Choose your option: ')))
    if start_choice == 1:

        def display_calendar():
            # Asking user for year and month to print calendar for that month.
            year = int(input("Provide year: "))
            month = int(input("Provide month: "))
            
            print(calendar.month(theyear = year, themonth = month, w = 5, l = 1))
            
        def add_task():
            # Asking user for year, month and day to add task.
            year = int(input("Provide year: "))
            month = int(input("Provide month: "))
            print(calendar.month(theyear = year, themonth = month, w = 5, l = 1))
            day = int(input("Provide day: "))
            
            # Task's content:
            task = input("Provide task's content: ")
            
            # Saving task to a file in YYYY-MM-DD: task format.
            with open("tasks.txt", "a") as f:
                f.write("{:04d}-{:02d}-{:02d}: {}\n".format(year, month, day, task))
            
            print("Task has been added.")

        def display_tasks():
            # Open file with tasks and print exercises for each month.
            year = int(input("Provide year: "))
            month = int(input("Provide month: "))
            
            with open("tasks.txt", "r") as f:
                tasks = [line.strip() for line in f.readlines()]
            
            print("Tasks for {} {}: ".format(calendar.month_name[month], year))
            for task in tasks:
                date_of_task, content_of_file = task.split(": ")
                task_year, task_month, task_day = map(int, date_of_task.split("-"))
                if task_year == year and task_month == month:
                    print("{:2d}: {}".format(task_day, content_of_file))

        # Main loop of 
        while True:
            print("\nWhat do you want to make?")
            print("1. Print Calendar")
            print("2. Add a task")
            print("3. Print tasks")
            print("4. Exit")
            
            choice = input("Choose your option (1-4): ")
            
            if choice == "1":
                display_calendar()
            elif choice == "2":
                add_task()
            elif choice == "3":
                display_tasks()
            elif choice == "4":
                break
            else:
                print("Incorrect choice. Please choose again.")
 
    elif start_choice == 2:

        # Excel file's name
        filename = "finanse.xlsx"

        # Opening or creating new file (if exists)
        try:
            wb = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            wb = openpyxl.Workbook()

        # Select worksheet "Incomes and expenses" or create new one
        if "Incomes and expenses" in wb.sheetnames:
            sheet = wb["Incomes and expenses"]
        else:
            sheet = wb.active
            sheet.title = "Incomes and expenses"

        # Display main menu
        while True:
            print("Choose your option:")
            print("1. Add a new expense")
            print("2. Add a new income")
            print("3. Display data")
            print("4. End the program")

            # Option selection
            choice = input("I am choosing: ")

            # Adding a new expense
            if choice == "1":
                amount = float(input("Provide expense amount: "))
                description = input("Enter a description of the expense: ")
                date = input("Provide date (YYYY-MM-DD): ")
                sheet.append(["Expense:", amount, description, date])
                print("A new expense has been added.")

            # Adding a new income
            elif choice == "2":
                amount = float(input("Provide income amount: "))
                description = input("Enter a description of the expense: ")
                date = input("Provide date (YYYY-MM-DD): ")
                sheet.append(["Income:", amount, description, date])
                print("A new income has been added.")
            
            # Printing data from excel file
            elif choice == "3":
                for row in sheet.iter_rows():
                    counter = 0
                    for cell in row:
                        print(cell.value, end=" | ")
                        counter += 1
                        if counter == 4:
                            print('\n')
                            counter = 0

            elif choice == "4":
                wb.save(filename)
                print("All data has been saved.")
                break
            
            else:
                print("Incorrect choice.")
    elif start_choice == 3:
        break
